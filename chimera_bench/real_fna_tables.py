from __future__ import annotations

import csv
import gzip
import hashlib
import re
from pathlib import Path
from typing import Iterable


COHORT_NAME_MAP = {
    "2015 Yu et al": "2015 Yu et al",
    "2018 Wirbel et al": "2018 Wirbel et al",
    "2019 Yachida et al": "YachidaS_2019",
}
SIGNAL_TOOL_NAMES = {
    "chimera_default": "Chimera",
    "centrifuger_default": "Centrifuger",
    "kraken2_safe_lf01": "Kraken2_LF01",
    "sylph_default": "sylph",
}


def read_tsv(paths: Iterable[Path] | Path) -> list[dict[str, str]]:
    if isinstance(paths, Path):
        paths = [paths]
    rows: list[dict[str, str]] = []
    for path in paths:
        with path.open(newline="") as handle:
            rows.extend(csv.DictReader(handle, delimiter="\t"))
    return rows


def _write_tsv(path: Path, rows: list[dict], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, delimiter="\t", fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def build_sample_manifest_rows(
    metadata_rows: list[dict[str, object]],
    source_rows: list[dict[str, str]],
    signal_rows: list[dict[str, str]],
) -> list[dict]:
    depth_rows = {
        row["sample"]: row for row in signal_rows if row["tool"] == "chimera_default"
    }
    if len(depth_rows) != len(source_rows):
        raise ValueError(
            f"sample/depth coverage mismatch: source={len(source_rows)} depth={len(depth_rows)}"
        )

    rows = []
    matched_paper_keys = set()
    for source in source_rows:
        sample = source["sample"]
        cohort = source["cohort"]
        candidates = [
            metadata
            for metadata in metadata_rows
            if metadata["cohort"] == cohort
            and f"_{metadata['paper_sample']}_" in f"_{sample}_"
        ]
        if len(candidates) != 1:
            raise ValueError(f"expected one paper metadata row for {cohort}/{sample}, found {len(candidates)}")
        metadata = candidates[0]
        paper_key = (cohort, str(metadata["paper_sample"]))
        if paper_key in matched_paper_keys:
            raise ValueError(f"paper sample matched twice: {paper_key}")
        matched_paper_keys.add(paper_key)
        if source["condition"].lower() != str(metadata["condition"]).lower():
            raise ValueError(f"condition mismatch for {sample}")

        signal = depth_rows[sample]
        c1 = float(source["paper_fna_c1_pct"])
        c2 = float(source["paper_fna_c2_pct"])
        total = float(source.get("paper_fna_total_pct") or c1 + c2)
        if abs(c1 - float(metadata["published_fna_c1_percent"])) > 1e-9:
            raise ValueError(f"published C1 mismatch for {sample}")
        if abs(c2 - float(metadata["published_fna_c2_percent"])) > 1e-9:
            raise ValueError(f"published C2 mismatch for {sample}")
        rows.append(
            {
                "sample": sample,
                "paper_sample_id": str(metadata["paper_sample"]),
                "run_accession": source["run_accession"],
                "cohort": cohort,
                "condition": source["condition"].lower(),
                "role": source["role"],
                "age": float(metadata["age"]),
                "sex": str(metadata["sex"]).lower(),
                "bmi": float(metadata["bmi"]),
                "source_reads": int(float(metadata["source_reads"])),
                "input_reads_used": int(float(signal["total_or_detected"])),
                "input_selection": "R1_head3m_or_all_available_if_short",
                "paper_fna_c1_percent": c1,
                "paper_fna_c2_percent": c2,
                "paper_fna_total_percent": total,
                "expected_c2_reads_at_depth": float(signal["expected_c2_reads_at_depth"]),
                "input_qc": "pass",
            }
        )
    if len(matched_paper_keys) != len(metadata_rows):
        raise ValueError(
            f"paper metadata coverage mismatch: matched={len(matched_paper_keys)} metadata={len(metadata_rows)}"
        )
    return rows


def build_signal_rows(
    signal_rows: list[dict[str, str]], manifest_by_sample: dict[str, dict]
) -> list[dict]:
    rows = []
    for source in signal_rows:
        sample = source["sample"]
        if sample not in manifest_by_sample:
            raise ValueError(f"signal sample missing from manifest: {sample}")
        metadata = manifest_by_sample[sample]
        tool = SIGNAL_TOOL_NAMES.get(source["tool"])
        if tool is None:
            raise ValueError(f"unsupported Fna signal tool: {source['tool']}")
        rows.append(
            {
                "sample": sample,
                "cohort": metadata["cohort"],
                "condition": metadata["condition"],
                "role": metadata["role"],
                "tool": tool,
                "signal_unit": source["unit"],
                "input_reads_used": metadata["input_reads_used"],
                "paper_fna_c2_percent": metadata["paper_fna_c2_percent"],
                "expected_c2_reads_at_depth": float(source["expected_c2_reads_at_depth"]),
                "fna_c2_signal": float(source["Fna_C2_signal"]),
                "fna_c1_signal": float(source["Fna_C1_signal"]),
                "non_c1c2_signal": float(source["non_C1C2_signal"]),
                "other_or_unmapped_signal": float(source["other_or_unmapped_signal"]),
            }
        )
    return sorted(rows, key=lambda row: (str(row["sample"]), str(row["tool"])))


def build_audit_rows(
    tool: str,
    audit_rows: list[dict[str, str]],
    manifest_by_sample: dict[str, dict],
) -> list[dict]:
    rows = []
    for source in audit_rows:
        sample = source["sample"]
        if sample not in manifest_by_sample:
            raise ValueError(f"audit sample missing from manifest: {sample}")
        metadata = manifest_by_sample[sample]
        input_reads = int(metadata["input_reads_used"])
        expected_c2_reads = float(
            metadata.get("expected_c2_reads_at_depth")
            if metadata.get("expected_c2_reads_at_depth") is not None
            else float(metadata["paper_fna_c2_percent"]) * input_reads / 100.0
        )
        candidate = int(source["selected_reads"])
        strict = int(source["strict_c2_best_reads"])
        tied = int(source["c2_tied_best_reads"])
        supported = int(source["c2_supported_reads"])
        non_c2 = int(source["non_c2_best_reads"])
        unmapped = int(source["unmapped_reads"])
        if strict + tied != supported:
            raise ValueError(f"C2 support count mismatch for {tool}/{sample}")
        if strict + tied + non_c2 + unmapped != candidate:
            raise ValueError(f"audit count closure failed for {tool}/{sample}")
        if int(source.get("missing_reads") or 0) != 0:
            raise ValueError(f"candidate FASTQ coverage failed for {tool}/{sample}")
        rows.append(
            {
                "sample": sample,
                "cohort": metadata["cohort"],
                "condition": metadata["condition"],
                "role": metadata["role"],
                "tool": tool,
                "input_reads_used": input_reads,
                "expected_c2_reads_at_depth": expected_c2_reads,
                "candidate_reads": candidate,
                "strict_c2_best_reads": strict,
                "c2_tied_best_reads": tied,
                "c2_supported_reads": supported,
                "non_c2_best_reads": non_c2,
                "unmapped_reads": unmapped,
                "strict_c2_best_rate": float(source["strict_c2_best_rate"]),
                "c2_tied_best_rate": float(source["c2_tied_best_rate"]),
                "c2_supported_rate": float(source["c2_supported_rate"]),
                "non_c2_best_rate": float(source["non_c2_best_rate"]),
                "unmapped_rate": float(source["unmapped_rate"]),
                "candidate_reads_per_million": candidate / input_reads * 1_000_000.0,
                "strict_c2_best_reads_per_million": strict / input_reads * 1_000_000.0,
                "c2_supported_reads_per_million": supported / input_reads * 1_000_000.0,
            }
        )
    return rows


def _fasta_stats(path: Path) -> tuple[int, int]:
    opener = gzip.open if path.suffix == ".gz" else open
    records = 0
    bases = 0
    with opener(path, "rt") as handle:
        for line in handle:
            if line.startswith(">"):
                records += 1
            else:
                bases += len(line.strip())
    return records, bases


def build_reference_rows(reference_rows: list[dict[str, str]]) -> list[dict]:
    rows = []
    for source in reference_rows:
        path = Path(source["path"])
        records, bases = _fasta_stats(path)
        accession_match = re.search(r"(GC[AF]_\d+\.\d+)", source["reference_name"])
        rows.append(
            {
                "reference_name": source["reference_name"],
                "accession": accession_match.group(1) if accession_match else "",
                "taxid": int(source["taxid"]),
                "paper_clade": source["paper_clade_projection"],
                "source_kind": "RefSeq" if accession_match else "paper_aggregate",
                "sequence_records": records,
                "sequence_bases": bases,
                "sha256": hashlib.sha256(path.read_bytes()).hexdigest(),
            }
        )
    return rows


def load_paper_metadata(path: Path) -> list[dict[str, object]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["ST23"]
    iterator = sheet.iter_rows(values_only=True)
    headers = next(iterator)
    positions = {str(value): index for index, value in enumerate(headers)}
    required = (
        "Cohort",
        "Sample ID",
        "Condition",
        "Age",
        "BMI",
        "Sex",
        "Number of Reads",
        "Percent Relative Abundance Fna C1",
        "Percent Relative Abundance Fna C2",
    )
    missing = [field for field in required if field not in positions]
    if missing:
        raise ValueError(f"ST23 missing columns: {missing}")
    rows = []
    for values in iterator:
        source_cohort = str(values[positions["Cohort"]])
        if source_cohort not in COHORT_NAME_MAP:
            continue
        rows.append(
            {
                "cohort": COHORT_NAME_MAP[source_cohort],
                "paper_sample": str(values[positions["Sample ID"]]),
                "condition": str(values[positions["Condition"]]).lower(),
                "age": float(values[positions["Age"]]),
                "bmi": float(values[positions["BMI"]]),
                "sex": str(values[positions["Sex"]]).lower(),
                "source_reads": float(values[positions["Number of Reads"]]),
                "published_fna_c1_percent": float(
                    values[positions["Percent Relative Abundance Fna C1"]]
                ),
                "published_fna_c2_percent": float(
                    values[positions["Percent Relative Abundance Fna C2"]]
                ),
            }
        )
    return rows


def write_fna_paper_tables(
    *,
    supplement_xlsx: Path,
    source_manifest_paths: list[Path],
    signal_paths: list[Path],
    audit_paths: list[tuple[str, Path]],
    reference_audit_path: Path,
    out_dir: Path,
) -> dict[str, int]:
    metadata_rows = load_paper_metadata(supplement_xlsx)
    source_rows = read_tsv(source_manifest_paths)
    source_signals = read_tsv(signal_paths)
    manifest_rows = build_sample_manifest_rows(metadata_rows, source_rows, source_signals)
    manifest_by_sample = {row["sample"]: row for row in manifest_rows}
    signal_rows = build_signal_rows(source_signals, manifest_by_sample)
    audit_rows = []
    for tool, path in audit_paths:
        audit_rows.extend(build_audit_rows(tool, read_tsv(path), manifest_by_sample))
    audit_rows.sort(key=lambda row: (str(row["sample"]), str(row["tool"])))
    reference_rows = build_reference_rows(read_tsv(reference_audit_path))

    _write_tsv(out_dir / "sample_manifest.tsv", manifest_rows, list(manifest_rows[0]))
    _write_tsv(out_dir / "sample_level_signals.tsv", signal_rows, list(signal_rows[0]))
    _write_tsv(out_dir / "read_audit_sample_metrics.tsv", audit_rows, list(audit_rows[0]))
    _write_tsv(out_dir / "reference_manifest.tsv", reference_rows, list(reference_rows[0]))
    return {
        "samples": len(manifest_rows),
        "signals": len(signal_rows),
        "audits": len(audit_rows),
        "references": len(reference_rows),
    }
