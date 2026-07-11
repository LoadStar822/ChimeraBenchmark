#!/usr/bin/env python3
from __future__ import annotations

import argparse
import bisect
import csv
import math
import random
import zlib
from pathlib import Path
from typing import Iterable


COHORT_NAME_MAP = {
    "2015 Yu et al": "2015 Yu et al",
    "2018 Wirbel et al": "2018 Wirbel et al",
    "2019 Yachida et al": "YachidaS_2019",
}
COHORT_ORDER = list(COHORT_NAME_MAP.values())
DEFAULT_INPUT_READS = 3_000_000
NORMAL_975 = 1.959963984540054


def auc_from_groups(positive: list[float], negative: list[float]) -> float:
    if not positive or not negative:
        raise ValueError("AUC requires non-empty positive and negative groups")

    ordered_negative = sorted(negative)
    wins = 0.0
    for pos in positive:
        lower = bisect.bisect_left(ordered_negative, pos)
        upper = bisect.bisect_right(ordered_negative, pos)
        wins += lower + 0.5 * (upper - lower)
    return wins / (len(positive) * len(negative))


def _invert_matrix(matrix: list[list[float]]) -> list[list[float]]:
    size = len(matrix)
    if size == 0 or any(len(row) != size for row in matrix):
        raise ValueError("matrix must be non-empty and square")

    augmented = [
        [float(value) for value in row]
        + [1.0 if row_index == column_index else 0.0 for column_index in range(size)]
        for row_index, row in enumerate(matrix)
    ]
    for column in range(size):
        pivot = max(range(column, size), key=lambda row: abs(augmented[row][column]))
        if abs(augmented[pivot][column]) < 1e-12:
            raise ValueError("design matrix is singular")
        augmented[column], augmented[pivot] = augmented[pivot], augmented[column]

        scale = augmented[column][column]
        augmented[column] = [value / scale for value in augmented[column]]
        for row in range(size):
            if row == column:
                continue
            factor = augmented[row][column]
            augmented[row] = [
                value - factor * pivot_value
                for value, pivot_value in zip(augmented[row], augmented[column])
            ]
    return [row[size:] for row in augmented]


def _beta_continued_fraction(a: float, b: float, x: float) -> float:
    maximum_iterations = 200
    epsilon = 3e-14
    floor = 1e-300
    qab = a + b
    qap = a + 1.0
    qam = a - 1.0
    c = 1.0
    d = 1.0 - qab * x / qap
    if abs(d) < floor:
        d = floor
    d = 1.0 / d
    result = d
    for iteration in range(1, maximum_iterations + 1):
        even = 2 * iteration
        coefficient = iteration * (b - iteration) * x / ((qam + even) * (a + even))
        d = 1.0 + coefficient * d
        if abs(d) < floor:
            d = floor
        c = 1.0 + coefficient / c
        if abs(c) < floor:
            c = floor
        d = 1.0 / d
        result *= d * c

        coefficient = -(a + iteration) * (qab + iteration) * x / (
            (a + even) * (qap + even)
        )
        d = 1.0 + coefficient * d
        if abs(d) < floor:
            d = floor
        c = 1.0 + coefficient / c
        if abs(c) < floor:
            c = floor
        d = 1.0 / d
        delta = d * c
        result *= delta
        if abs(delta - 1.0) < epsilon:
            return result
    raise ArithmeticError("incomplete beta continued fraction did not converge")


def _regularized_incomplete_beta(a: float, b: float, x: float) -> float:
    if a <= 0 or b <= 0 or not 0.0 <= x <= 1.0:
        raise ValueError("invalid incomplete beta arguments")
    if x == 0.0:
        return 0.0
    if x == 1.0:
        return 1.0
    scale = math.exp(
        math.lgamma(a + b)
        - math.lgamma(a)
        - math.lgamma(b)
        + a * math.log(x)
        + b * math.log1p(-x)
    )
    if x < (a + 1.0) / (a + b + 2.0):
        return scale * _beta_continued_fraction(a, b, x) / a
    return 1.0 - scale * _beta_continued_fraction(b, a, 1.0 - x) / b


def two_sided_t_p_value(statistic: float, *, residual_df: int) -> float:
    if residual_df <= 0:
        raise ValueError("t test requires positive residual degrees of freedom")
    if math.isinf(statistic):
        return 0.0
    x = residual_df / (residual_df + statistic * statistic)
    return _regularized_incomplete_beta(residual_df / 2.0, 0.5, x)


def fit_ols(
    response: list[float],
    design: list[list[float]],
    *,
    coefficient_index: int,
) -> dict[str, float | int]:
    if not response or len(response) != len(design):
        raise ValueError("response and design must have the same non-zero length")
    width = len(design[0])
    if width == 0 or any(len(row) != width for row in design):
        raise ValueError("design rows must have a consistent non-zero width")
    if not 0 <= coefficient_index < width:
        raise ValueError("coefficient index is outside the design matrix")
    residual_df = len(response) - width
    if residual_df <= 0:
        raise ValueError("OLS requires more observations than coefficients")

    xtx = [
        [sum(row[left] * row[right] for row in design) for right in range(width)]
        for left in range(width)
    ]
    xty = [
        sum(row[column] * value for row, value in zip(design, response))
        for column in range(width)
    ]
    inverse = _invert_matrix(xtx)
    coefficients = [sum(value * target for value, target in zip(row, xty)) for row in inverse]
    residuals = [
        observed - sum(value * coefficient for value, coefficient in zip(row, coefficients))
        for observed, row in zip(response, design)
    ]
    residual_variance = sum(value * value for value in residuals) / residual_df
    standard_error = math.sqrt(residual_variance * inverse[coefficient_index][coefficient_index])
    coefficient = coefficients[coefficient_index]
    statistic = coefficient / standard_error if standard_error else math.copysign(math.inf, coefficient)
    p_value = two_sided_t_p_value(statistic, residual_df=residual_df)
    return {
        "coefficient": coefficient,
        "standard_error": standard_error,
        "statistic": statistic,
        "p_value": p_value,
        "residual_df": residual_df,
    }


def paper_standardized_effect(
    *,
    statistic: float,
    n_case: int,
    n_control: int,
) -> tuple[float, float]:
    total = n_case + n_control
    if n_case <= 0 or n_control <= 0 or total <= 3:
        raise ValueError("standardized effect requires both groups and at least four samples")
    effect = statistic * total / math.sqrt(n_case * n_control * (total - 2))
    standard_error = math.sqrt(
        ((total - 1) / (total - 3))
        * (4 / total)
        * (1 + effect * effect / 8)
    )
    return effect, standard_error


def _weighted_effect(
    effects: list[float],
    variances: list[float],
    tau_squared: float,
) -> tuple[float, float, float]:
    weights = [1.0 / (variance + tau_squared) for variance in variances]
    weight_sum = sum(weights)
    effect = sum(weight * value for weight, value in zip(weights, effects)) / weight_sum
    q_value = sum(weight * (value - effect) ** 2 for weight, value in zip(weights, effects))
    return effect, weight_sum, q_value


def paule_mandel_meta(effects: list[float], variances: list[float]) -> dict[str, float | int]:
    if len(effects) < 2 or len(effects) != len(variances):
        raise ValueError("meta-analysis requires matching effects and variances from at least two cohorts")
    if any(variance <= 0 for variance in variances):
        raise ValueError("meta-analysis variances must be positive")

    degrees_of_freedom = len(effects) - 1
    _, _, q_at_zero = _weighted_effect(effects, variances, 0.0)
    if q_at_zero <= degrees_of_freedom:
        tau_squared = 0.0
    else:
        low = 0.0
        high = max(variances)
        while _weighted_effect(effects, variances, high)[2] > degrees_of_freedom:
            high *= 2.0
        for _ in range(100):
            middle = (low + high) / 2.0
            if _weighted_effect(effects, variances, middle)[2] > degrees_of_freedom:
                low = middle
            else:
                high = middle
        tau_squared = (low + high) / 2.0

    effect, weight_sum, _ = _weighted_effect(effects, variances, tau_squared)
    standard_error = math.sqrt(1.0 / weight_sum)
    statistic = effect / standard_error
    i_squared = max(0.0, (q_at_zero - degrees_of_freedom) / q_at_zero) if q_at_zero else 0.0
    return {
        "effect": effect,
        "standard_error": standard_error,
        "ci_low": effect - NORMAL_975 * standard_error,
        "ci_high": effect + NORMAL_975 * standard_error,
        "wald_p_value": math.erfc(abs(statistic) / math.sqrt(2.0)),
        "tau_squared": tau_squared,
        "q_value": q_at_zero,
        "i_squared": i_squared,
        "cohorts": len(effects),
    }


def _quantile(values: list[float], probability: float) -> float:
    ordered = sorted(values)
    if not ordered:
        raise ValueError("quantile requires at least one value")
    position = probability * (len(ordered) - 1)
    lower = int(math.floor(position))
    upper = int(math.ceil(position))
    if lower == upper:
        return ordered[lower]
    fraction = position - lower
    return ordered[lower] * (1.0 - fraction) + ordered[upper] * fraction


def stratified_bootstrap_auc(
    positive: list[float],
    negative: list[float],
    *,
    iterations: int,
    seed: int,
) -> dict[str, float | int]:
    if iterations <= 0:
        raise ValueError("bootstrap iterations must be positive")
    observed = auc_from_groups(positive, negative)
    generator = random.Random(seed)
    bootstrap = []
    for _ in range(iterations):
        sampled_positive = [generator.choice(positive) for _ in positive]
        sampled_negative = [generator.choice(negative) for _ in negative]
        bootstrap.append(auc_from_groups(sampled_positive, sampled_negative))
    return {
        "auc": observed,
        "ci_low": _quantile(bootstrap, 0.025),
        "ci_high": _quantile(bootstrap, 0.975),
        "iterations": iterations,
        "seed": seed,
    }


def match_metadata_rows(
    metadata_rows: list[dict[str, object]],
    sample_rows: list[dict[str, str]],
) -> list[dict[str, object]]:
    by_cohort: dict[str, list[dict[str, str]]] = {}
    for row in sample_rows:
        by_cohort.setdefault(row["cohort"], []).append(row)

    matched = []
    for metadata in metadata_rows:
        cohort = str(metadata["cohort"])
        paper_sample = str(metadata["paper_sample"])
        frozen_sample = str(metadata.get("sample") or "")
        if frozen_sample:
            candidates = [
                row
                for row in by_cohort.get(cohort, [])
                if row["sample"] == frozen_sample
            ]
        else:
            candidates = [
                row
                for row in by_cohort.get(cohort, [])
                if f"_{paper_sample}_" in f"_{row['sample']}_"
            ]
        if len(candidates) != 1:
            raise ValueError(
                f"expected one unique audit sample for {cohort}/{paper_sample}, found {len(candidates)}"
            )
        candidate = candidates[0]
        paper_condition = str(metadata.get("condition", "")).lower()
        audit_condition = str(candidate.get("condition", "")).lower()
        if paper_condition and audit_condition and paper_condition != audit_condition:
            raise ValueError(
                f"condition mismatch for {cohort}/{paper_sample}: "
                f"paper={paper_condition}, audit={audit_condition}"
            )
        matched.append({**candidate, **metadata, "audit_condition": audit_condition})
    return matched


def read_tsv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def load_frozen_metadata(path: Path) -> list[dict[str, object]]:
    required = {
        "sample",
        "paper_sample_id",
        "cohort",
        "condition",
        "age",
        "bmi",
        "sex",
        "source_reads",
        "paper_fna_c1_percent",
        "paper_fna_c2_percent",
    }
    rows = read_tsv(path)
    missing = required - set(rows[0] if rows else ())
    if missing:
        raise ValueError(f"frozen sample manifest is missing columns: {sorted(missing)}")
    return [
        {
            "sample": row["sample"],
            "cohort": row["cohort"],
            "paper_sample": row["paper_sample_id"],
            "condition": row["condition"].lower(),
            "age": float(row["age"]),
            "bmi": float(row["bmi"]),
            "sex": row["sex"].lower(),
            "source_reads": float(row["source_reads"]),
            "published_fna_c1_percent": float(row["paper_fna_c1_percent"]),
            "published_fna_c2_percent": float(row["paper_fna_c2_percent"]),
        }
        for row in rows
    ]


def load_combined_sample_tables(paths: list[Path]) -> dict[str, list[dict[str, str]]]:
    grouped: dict[str, list[dict[str, str]]] = {}
    seen: set[tuple[str, str]] = set()
    for path in paths:
        for row in read_tsv(path):
            if not row.get("tool") or not row.get("sample"):
                raise ValueError(f"combined sample table lacks tool/sample values: {path}")
            key = row["tool"], row["sample"]
            if key in seen:
                raise ValueError(f"duplicate tool/sample row: {key}")
            seen.add(key)
            grouped.setdefault(row["tool"], []).append(row)
    return grouped


def write_tsv(path: Path, rows: Iterable[dict[str, object]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, delimiter="\t", fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def load_paper_metadata(path: Path) -> list[dict[str, object]]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError("openpyxl is required to read the source supplementary workbook") from error

    workbook = load_workbook(path, read_only=True, data_only=True)
    if "ST23" not in workbook.sheetnames:
        raise ValueError(f"supplementary workbook has no ST23 sheet: {path}")
    sheet = workbook["ST23"]
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows)
    positions = {str(value): index for index, value in enumerate(headers)}
    required = [
        "Cohort",
        "Sample ID",
        "Condition",
        "Age",
        "BMI",
        "Sex",
        "Number of Reads",
        "Percent Relative Abundance Fna C1",
        "Percent Relative Abundance Fna C2",
    ]
    missing = [field for field in required if field not in positions]
    if missing:
        raise ValueError(f"ST23 is missing required columns: {missing}")

    metadata = []
    for row in rows:
        source_cohort = str(row[positions["Cohort"]])
        if source_cohort not in COHORT_NAME_MAP:
            continue
        metadata.append(
            {
                "cohort": COHORT_NAME_MAP[source_cohort],
                "paper_sample": str(row[positions["Sample ID"]]),
                "condition": str(row[positions["Condition"]]).lower(),
                "age": float(row[positions["Age"]]),
                "bmi": float(row[positions["BMI"]]),
                "sex": str(row[positions["Sex"]]).lower(),
                "source_reads": float(row[positions["Number of Reads"]]),
                "published_fna_c1_percent": float(
                    row[positions["Percent Relative Abundance Fna C1"]]
                ),
                "published_fna_c2_percent": float(
                    row[positions["Percent Relative Abundance Fna C2"]]
                ),
            }
        )
    return metadata


def load_input_read_counts(paths: list[Path]) -> dict[str, int]:
    counts: dict[str, int] = {}
    for path in paths:
        for row in read_tsv(path):
            sample = row["sample"]
            reads = int(row["input_reads_used"])
            if sample in counts and counts[sample] != reads:
                raise ValueError(f"conflicting input read counts for {sample}")
            counts[sample] = reads
    return counts


def _parse_named_path(value: str) -> tuple[str, Path]:
    if "=" not in value:
        raise argparse.ArgumentTypeError("expected TOOL=PATH")
    name, raw_path = value.split("=", 1)
    if not name or not raw_path:
        raise argparse.ArgumentTypeError("expected non-empty TOOL=PATH")
    return name, Path(raw_path)


def _stable_seed(base_seed: int, *parts: str) -> int:
    token = "\x1f".join(parts).encode()
    return (base_seed + zlib.crc32(token)) % (2**32)


def _format_float(value: float) -> str:
    return f"{value:.6f}"


def _format_p_value(value: float) -> str:
    return f"{value:.6g}"


def _analyze_prepared_metrics(
    tool: str,
    prepared: list[dict[str, object]],
    metric_specs: list[tuple[str, str]],
    *,
    bootstrap_iterations: int,
    seed: int,
) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    cohort_rows: list[dict[str, object]] = []
    meta_rows: list[dict[str, object]] = []
    for metric, model_metric in metric_specs:
        effects = []
        variances = []
        total_case = 0
        total_control = 0
        for cohort in COHORT_ORDER:
            group = [row for row in prepared if row["cohort"] == cohort]
            if not group:
                raise ValueError(f"{tool}: no matched rows for cohort {cohort}")
            positive = [float(row[metric]) for row in group if row["case"] == 1]
            negative = [float(row[metric]) for row in group if row["case"] == 0]
            bootstrap = stratified_bootstrap_auc(
                positive,
                negative,
                iterations=bootstrap_iterations,
                seed=_stable_seed(seed, tool, metric, cohort),
            )
            response = [
                math.asin(math.sqrt(min(1.0, max(0.0, float(row[model_metric])))))
                for row in group
            ]
            design = [
                [
                    1.0,
                    float(row["case"]),
                    float(row["male"]),
                    float(row["age"]),
                    float(row["bmi"]),
                    float(row["source_reads"]) / 10_000_000.0,
                ]
                for row in group
            ]
            model = fit_ols(response, design, coefficient_index=1)
            effect, effect_se = paper_standardized_effect(
                statistic=float(model["statistic"]),
                n_case=len(positive),
                n_control=len(negative),
            )
            effects.append(effect)
            variances.append(effect_se * effect_se)
            total_case += len(positive)
            total_control += len(negative)
            cohort_rows.append(
                {
                    "tool": tool,
                    "metric": metric,
                    "cohort": cohort,
                    "n_case": len(positive),
                    "n_control": len(negative),
                    "case_mean": _format_float(sum(positive) / len(positive)),
                    "control_mean": _format_float(sum(negative) / len(negative)),
                    "auc_crc_vs_control": _format_float(float(bootstrap["auc"])),
                    "auc_ci_low": _format_float(float(bootstrap["ci_low"])),
                    "auc_ci_high": _format_float(float(bootstrap["ci_high"])),
                    "adjusted_condition_beta": _format_float(float(model["coefficient"])),
                    "adjusted_condition_se": _format_float(float(model["standard_error"])),
                    "adjusted_wald_statistic": _format_float(float(model["statistic"])),
                    "adjusted_p_value": _format_p_value(float(model["p_value"])),
                    "adjusted_standardized_effect": _format_float(effect),
                    "adjusted_effect_se": _format_float(effect_se),
                }
            )
        meta = paule_mandel_meta(effects, variances)
        meta_rows.append(
            {
                "tool": tool,
                "metric": metric,
                "cohorts": int(meta["cohorts"]),
                "matched_samples": len(prepared),
                "n_case": total_case,
                "n_control": total_control,
                "random_effect_standardized_effect": _format_float(float(meta["effect"])),
                "effect_ci_low": _format_float(float(meta["ci_low"])),
                "effect_ci_high": _format_float(float(meta["ci_high"])),
                "wald_p_value": _format_p_value(float(meta["wald_p_value"])),
                "tau_squared": _format_float(float(meta["tau_squared"])),
                "q_value": _format_float(float(meta["q_value"])),
                "i_squared": _format_float(float(meta["i_squared"])),
            }
        )
    return cohort_rows, meta_rows


def analyze_tool(
    tool: str,
    metadata_rows: list[dict[str, object]],
    sample_rows: list[dict[str, str]],
    input_read_counts: dict[str, int],
    *,
    bootstrap_iterations: int,
    seed: int,
) -> tuple[list[dict[str, object]], list[dict[str, object]], dict[str, object]]:
    matched = sorted(
        match_metadata_rows(metadata_rows, sample_rows),
        key=lambda row: (str(row["cohort"]), str(row["paper_sample"])),
    )
    if len(matched) != len(sample_rows):
        raise ValueError(
            f"{tool}: matched {len(matched)} paper rows but audit table has {len(sample_rows)} rows"
        )
    matched_names = [str(row["sample"]) for row in matched]
    if len(set(matched_names)) != len(matched_names):
        raise ValueError(f"{tool}: an audit sample matched more than one paper row")

    prepared = []
    for row in matched:
        condition = str(row["condition"]).lower()
        audit_condition = str(row.get("audit_condition", "")).lower()
        if condition not in {"disease", "control"} or audit_condition != condition:
            raise ValueError(f"{tool}: condition mismatch for {row['sample']}")
        sex = str(row["sex"]).lower()
        if sex not in {"female", "male"}:
            raise ValueError(f"{tool}: unsupported sex value {sex!r} for {row['sample']}")
        input_reads = input_read_counts.get(
            str(row["sample"]),
            int(float(str(row.get("input_reads_used") or DEFAULT_INPUT_READS))),
        )
        if input_reads <= 0:
            raise ValueError(f"{tool}: non-positive input read count for {row['sample']}")
        strict_reads = int(str(row["strict_c2_best_reads"]))
        strict_rate = float(str(row["strict_c2_best_rate"]))
        prepared.append(
            {
                **row,
                "case": int(condition == "disease"),
                "male": int(sex == "male"),
                "input_reads": input_reads,
                "strict_c2_best_rate": strict_rate,
                "strict_c2_best_rate_fraction": strict_rate,
                "strict_c2_best_reads_per_million": strict_reads / input_reads * 1_000_000.0,
                "strict_c2_best_reads_per_million_fraction": strict_reads / input_reads,
            }
        )

    metric_specs = [
        ("strict_c2_best_reads_per_million", "strict_c2_best_reads_per_million_fraction"),
        ("strict_c2_best_rate", "strict_c2_best_rate_fraction"),
    ]
    cohort_rows, meta_rows = _analyze_prepared_metrics(
        tool,
        prepared,
        metric_specs,
        bootstrap_iterations=bootstrap_iterations,
        seed=seed,
    )
    qc = {
        "tool": tool,
        "paper_metadata_samples": len(metadata_rows),
        "audit_samples": len(sample_rows),
        "matched_samples": len(prepared),
        "unique_matched_samples": len(set(matched_names)),
        "status": "pass",
    }
    return cohort_rows, meta_rows, qc


def analyze_published_fna_c2(
    metadata_rows: list[dict[str, object]],
    *,
    bootstrap_iterations: int,
    seed: int,
) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    prepared: list[dict[str, object]] = []
    for row in sorted(
        metadata_rows,
        key=lambda item: (str(item["cohort"]), str(item["paper_sample"])),
    ):
        condition = str(row["condition"]).lower()
        sex = str(row["sex"]).lower()
        if condition not in {"disease", "control"}:
            raise ValueError(f"Published_Fna_C2: unsupported condition {condition!r}")
        if sex not in {"female", "male"}:
            raise ValueError(f"Published_Fna_C2: unsupported sex {sex!r}")
        percent = float(row["published_fna_c2_percent"])
        prepared.append(
            {
                **row,
                "case": int(condition == "disease"),
                "male": int(sex == "male"),
                "published_fna_c2_percent": percent,
                "published_fna_c2_fraction": percent / 100.0,
            }
        )
    return _analyze_prepared_metrics(
        "Published_Fna_C2",
        prepared,
        [("published_fna_c2_percent", "published_fna_c2_fraction")],
        bootstrap_iterations=bootstrap_iterations,
        seed=seed,
    )


COHORT_FIELDS = [
    "tool",
    "metric",
    "cohort",
    "n_case",
    "n_control",
    "case_mean",
    "control_mean",
    "auc_crc_vs_control",
    "auc_ci_low",
    "auc_ci_high",
    "adjusted_condition_beta",
    "adjusted_condition_se",
    "adjusted_wald_statistic",
    "adjusted_p_value",
    "adjusted_standardized_effect",
    "adjusted_effect_se",
]
META_FIELDS = [
    "tool",
    "metric",
    "cohorts",
    "matched_samples",
    "n_case",
    "n_control",
    "random_effect_standardized_effect",
    "effect_ci_low",
    "effect_ci_high",
    "wald_p_value",
    "tau_squared",
    "q_value",
    "i_squared",
]
QC_FIELDS = [
    "tool",
    "paper_metadata_samples",
    "audit_samples",
    "matched_samples",
    "unique_matched_samples",
    "status",
]


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(
        description="Test whether audited Fna C2 read evidence recovers CRC associations across cohorts."
    )
    metadata_group = parser.add_mutually_exclusive_group(required=True)
    metadata_group.add_argument("--supplement-xlsx", type=Path)
    metadata_group.add_argument("--metadata-table", type=Path)
    parser.add_argument("--sample-table", action="append", default=[], type=_parse_named_path)
    parser.add_argument("--combined-sample-table", action="append", default=[], type=Path)
    parser.add_argument("--input-qc", action="append", default=[], type=Path)
    parser.add_argument("--out-dir", required=True, type=Path)
    parser.add_argument("--bootstrap", type=int, default=2_000)
    parser.add_argument("--seed", type=int, default=20_260_710)
    args = parser.parse_args(argv)

    metadata_rows = (
        load_paper_metadata(args.supplement_xlsx)
        if args.supplement_xlsx
        else load_frozen_metadata(args.metadata_table)
    )
    input_read_counts = load_input_read_counts(args.input_qc)
    tool_tables: dict[str, list[Path]] = {}
    for tool, path in args.sample_table:
        tool_tables.setdefault(tool, []).append(path)

    combined_rows = load_combined_sample_tables(args.combined_sample_table)
    if not tool_tables and not combined_rows:
        parser.error("at least one --sample-table or --combined-sample-table is required")

    all_cohort_rows = []
    all_meta_rows = []
    all_qc_rows = []
    all_tools = sorted(set(tool_tables) | set(combined_rows))
    for tool in all_tools:
        sample_rows = list(combined_rows.get(tool, []))
        for path in tool_tables.get(tool, []):
            sample_rows.extend(read_tsv(path))
        cohort_rows, meta_rows, qc = analyze_tool(
            tool,
            metadata_rows,
            sample_rows,
            input_read_counts,
            bootstrap_iterations=args.bootstrap,
            seed=args.seed,
        )
        all_cohort_rows.extend(cohort_rows)
        all_meta_rows.extend(meta_rows)
        all_qc_rows.append(qc)

    paper_cohort_rows, paper_meta_rows = analyze_published_fna_c2(
        metadata_rows,
        bootstrap_iterations=args.bootstrap,
        seed=args.seed,
    )
    all_cohort_rows.extend(paper_cohort_rows)
    all_meta_rows.extend(paper_meta_rows)

    write_tsv(args.out_dir / "crc_association_cohort.tsv", all_cohort_rows, COHORT_FIELDS)
    write_tsv(args.out_dir / "crc_association_meta.tsv", all_meta_rows, META_FIELDS)
    write_tsv(args.out_dir / "crc_association_match_qc.tsv", all_qc_rows, QC_FIELDS)
    print(
        f"wrote CRC association results for {len(all_tools)} tools, the published Fna C2 reference, "
        f"and {len(metadata_rows)} samples to {args.out_dir}"
    )


if __name__ == "__main__":
    main()
